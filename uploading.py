import streamlit as st
import pandas as pd
import openpyxl

# Initialize global variables
sold_df = None
unsold_df = None


def process_excel(uploaded_file):
    global sold_df, unsold_df  # Declare global variables

    # Open the Excel file using openpyxl
    wb = openpyxl.load_workbook(uploaded_file,
                                data_only=True)  # Use data_only=True to get cell values instead of formulas

    # Check if "Table C" exists in the workbook
    if "Table C" in wb.sheetnames:
        # Get the desired worksheet
        ws = wb["Table C"]

        # Initialize lists to store sold and unsold inventory
        sold_inventory = []
        unsold_inventory = []

        # Flag to determine whether to capture rows
        capture_sold = False
        capture_unsold = False

        # Iterate over rows in the worksheet
        for row in ws.iter_rows(values_only=True):
            # Check if the row contains "Sold Inventory"
            if "Sold Inventory" in str(row):
                capture_sold = True
                capture_unsold = False
                continue
            # Check if the row contains "Unsold Inventory"
            elif "Unsold Inventory" in str(row):
                capture_sold = False
                capture_unsold = True
                continue
            # Check if the row contains "Total" indicating the end of a table
            elif "Total" in str(row):
                continue

            # Append the row to the appropriate list after converting formulas to values
            if capture_sold:
                sold_inventory.append(row)
            elif capture_unsold:
                unsold_inventory.append(row)

        # Convert lists to DataFrames
        sold_df = pd.DataFrame(sold_inventory)
        unsold_df = pd.DataFrame(unsold_inventory)

        if sold_df is not None:
            # Extract first four columns from sold_df
            sold_df = sold_df.iloc[:, :4]

            # Convert "Carpet Area In Sq.Mtrs" column to float and round to 2 decimal places
            if "Carpet Area In Sq.Mtrs" in sold_df.columns:
                sold_df["Carpet Area In Sq.Mtrs"] = sold_df["Carpet Area In Sq.Mtrs"].apply(lambda x: float(x) if x is not None else None)
                sold_df["Carpet Area In Sq.Mtrs"] = sold_df["Carpet Area In Sq.Mtrs"].round(2)

        if unsold_df is not None:
            # Extract first four columns from unsold_df
            unsold_df = unsold_df.iloc[:, :4]

            # Convert "Carpet Area In Sq.Mtrs" column to float and round to 2 decimal places
            if "Carpet Area In Sq.Mtrs" in unsold_df.columns:
                unsold_df["Carpet Area In Sq.Mtrs"] = unsold_df["Carpet Area In Sq.Mtrs"].apply(lambda x: float(x) if x is not None else None)
                unsold_df["Carpet Area In Sq.Mtrs"] = unsold_df["Carpet Area In Sq.Mtrs"].round(2)

        return sold_df, unsold_df  # Return the modified dataframes


def get_building_details():
    global sold_df, unsold_df  # Access global variables

    uploaded_file = st.file_uploader("Dump your Form 3 here ", type=["xlsx", "xls", "xlsb"],
                                     help="Upload the new format form 3 excel.")
    if uploaded_file is not None:
        # Process the uploaded Excel file
        sold_df, unsold_df = process_excel(uploaded_file)

        if sold_df is not None and not sold_df.empty:
            st.write("Sold Inventory:")
            st.write(sold_df)
            st.write("___________________________________________")

        if unsold_df is not None and not unsold_df.empty:
            st.write("Unsold Inventory:")
            st.write(unsold_df)


# Run the app
if __name__ == "__main__":
    get_building_details()
